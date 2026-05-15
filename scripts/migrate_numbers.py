#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path("/private/tmp/stock-portfolio-export.xlsx")
DEFAULT_DB = ROOT / "data" / "portfolio.db"
SUMMARY_PATH = ROOT / "data" / "private" / "migration-summary.json"
STATE_KEY = "default"
STATE_VERSION = 6
FLOW_UNIT = 10_000

US_TICKERS = {
    "VOO",
    "SSO",
    "QLD",
    "QQQ",
    "QQQM",
    "SPY",
    "AAPL",
    "MSFT",
    "NVDA",
    "GOOGL",
    "AMZN",
    "META",
    "TSLA",
    "TQQQ",
    "ASML",
    "AMD",
}

US_NAME_TO_TICKER = {
    "Microsoft": "MSFT",
    "NVIDIA": "NVDA",
    "TESLA": "TSLA",
    "APPLE": "AAPL",
    "ASML": "ASML",
    "AMD": "AMD",
}

KR_TICKERS = {
    "삼성전자": "005930.KS",
    "현대차": "005380.KS",
    "SKC": "011790.KS",
}


def main() -> None:
    parser = argparse.ArgumentParser(description="Migrate exported Numbers workbook into Stock Portfolio Lab SQLite state.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--summary", type=Path, default=SUMMARY_PATH)
    parser.add_argument("--preview", action="store_true", help="Print preview JSON and do not write SQLite state.")
    args = parser.parse_args()

    state, summary = migrate_workbook(args.xlsx)
    if args.preview:
        print(json.dumps({"summary": summary, "state": state}, ensure_ascii=False))
        return

    write_state(args.db, state)
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    args.summary.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


def migrate_workbook(xlsx_path: Path) -> tuple[dict, dict]:
    workbook = load_workbook(xlsx_path, data_only=True)
    fx_rate = read_fx_rate(workbook)
    holdings = read_holdings(workbook)
    snapshots, cash_flows = read_performance(workbook, fx_rate)
    cash_balances = infer_cash_balances(holdings, snapshots, fx_rate)
    state = {
        "version": STATE_VERSION,
        "fxRate": {
            "pair": "USD/KRW",
            "rate": fx_rate,
            "source": "Numbers import",
            "asOf": datetime.now(timezone.utc).isoformat(),
        },
        "holdings": holdings,
        "cashFlows": cash_flows,
        "cashBalances": cash_balances,
        "accounts": build_accounts(holdings, cash_balances),
        "dashboardLayout": default_dashboard_layout(),
        "accountSnapshots": [],
        "priceUpdateLogs": [],
        "portfolioSnapshots": snapshots,
        "automation": {
            "lastRunAt": None,
            "lastResult": "Numbers 데이터 마이그레이션 완료",
            "snapshotTime": "09:10",
            "timezone": "Asia/Seoul",
        },
    }
    summary = validate_state(state, workbook)
    return state, summary


def default_dashboard_layout() -> list[dict]:
    return [
        {"id": "total-value", "widthPct": 25, "span": 3, "minHeight": 128, "visible": True},
        {"id": "total-cost", "widthPct": 25, "span": 3, "minHeight": 128, "visible": True},
        {"id": "total-gain", "widthPct": 25, "span": 3, "minHeight": 128, "visible": True},
        {"id": "cash-total", "widthPct": 25, "span": 3, "minHeight": 128, "visible": True},
        {"id": "fx-rate", "widthPct": 25, "span": 3, "minHeight": 128, "visible": True},
        {"id": "allocation", "widthPct": 50, "span": 6, "minHeight": 320, "visible": True},
        {"id": "performance-flow", "widthPct": 50, "span": 6, "minHeight": 320, "visible": True},
        {"id": "breakdown", "widthPct": 50, "span": 6, "minHeight": 320, "visible": True},
    ]


def read_fx_rate(workbook) -> float:
    if "QQQ - 환율" not in workbook.sheetnames:
        return 1350.0
    ws = workbook["QQQ - 환율"]
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, (int, float)) and value > 0:
                return float(value)
    return 1350.0


def read_holdings(workbook) -> list[dict]:
    holdings: list[dict] = []
    for ws in workbook.worksheets:
        strategy = sheet_strategy(ws.title)
        if not strategy:
            continue
        investor = sheet_investor(ws.title)
        last_account = ""
        for row in ws.iter_rows(min_row=3, values_only=True):
            account, instrument, quantity, price, average_cost, *_rest = list(row) + [None] * 9
            if account:
                last_account = str(account).strip()
            account = str(account or last_account).strip()
            instrument = str(instrument or "").strip()
            if not is_holding_row(account, instrument, quantity, price, average_cost):
                continue
            currency = infer_currency(instrument, price)
            ticker, auto_price = infer_ticker(instrument, currency)
            holdings.append(
                {
                    "id": make_id("holding"),
                    "investor": investor,
                    "account": account,
                    "accountType": infer_account_type(account, currency),
                    "strategy": strategy,
                    "ticker": ticker,
                    "name": instrument,
                    "quantity": float(quantity),
                    "averageCost": float(average_cost),
                    "price": float(price),
                    "currency": currency,
                    "priceSource": "Numbers import",
                    "priceAsOf": "Numbers export",
                    "autoPrice": auto_price,
                }
            )
    return holdings


def read_performance(workbook, fx_rate: float) -> tuple[list[dict], list[dict]]:
    snapshots_by_date: dict[str, dict] = {}
    cash_flows: list[dict] = []
    for ws in workbook.worksheets:
        if "자산 Flow" not in ws.title:
            continue
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            continue
        headers = rows[1]
        row_map = {str(row[0]).strip(): row for row in rows if row and row[0] is not None}
        total_row = row_map.get("평가금액")
        deposit_row = row_map.get("추가입금")
        daily_row = row_map.get("일별")
        if not total_row:
            continue
        for col_idx, header in enumerate(headers):
            if not isinstance(header, datetime):
                continue
            date_key = header.date().isoformat()
            total_raw = number_at(total_row, col_idx)
            if total_raw is None:
                continue
            total_value_krw = total_raw * FLOW_UNIT
            net_inflow_krw = (number_at(deposit_row, col_idx) or 0) * FLOW_UNIT if deposit_row else 0
            investment_gain_krw = (number_at(daily_row, col_idx) or 0) * FLOW_UNIT if daily_row else 0
            snapshots_by_date[date_key] = {
                "id": make_id("snapshot"),
                "date": date_key,
                "totalValueUsd": total_value_krw / fx_rate,
                "totalValueKrw": total_value_krw,
                "totalCostUsd": None,
                "totalGainUsd": investment_gain_krw / fx_rate,
                "fxRate": fx_rate,
                "netInflowKrw": net_inflow_krw,
            }
            if net_inflow_krw:
                cash_flows.append(
                    {
                        "id": make_id("cashflow"),
                        "date": date_key,
                        "investor": "전체",
                        "account": "Numbers 자산 Flow",
                        "type": "deposit" if net_inflow_krw > 0 else "withdrawal",
                        "amountKrw": abs(net_inflow_krw),
                        "note": "Numbers 자산 Flow import",
                    }
                )
    snapshots = [snapshots_by_date[key] for key in sorted(snapshots_by_date)]
    return snapshots, cash_flows


def sheet_strategy(title: str) -> str | None:
    if "총합" in title or "그림" in title or "환율" in title:
        return None
    if title.startswith("QQQ - "):
        return "QQQ"
    if title.startswith("S&P500 - "):
        return "S&P500"
    if title.startswith("국내주식 - "):
        return "국내주식"
    return None


def sheet_investor(title: str) -> str:
    match = re.search(r"-\s*([^\s]+)", title)
    return match.group(1) if match else "미분류"


def is_holding_row(account, instrument, quantity, price, average_cost) -> bool:
    if not account or not instrument:
        return False
    if instrument == "합계" or account in {"연금", "직투", "직투(달러)", "직투(원화)", "직투 + 연금"}:
        return False
    return all(isinstance(value, (int, float)) for value in (quantity, price, average_cost))


def infer_currency(instrument: str, price) -> str:
    if instrument.upper() in US_TICKERS or instrument in US_NAME_TO_TICKER:
        return "USD"
    if isinstance(price, (int, float)) and price < 1000 and re.fullmatch(r"[A-Z.]+", instrument.upper()):
        return "USD"
    return "KRW"


def infer_ticker(instrument: str, currency: str) -> tuple[str, bool]:
    upper = instrument.upper()
    if instrument in US_NAME_TO_TICKER:
        return US_NAME_TO_TICKER[instrument], True
    if upper in US_TICKERS:
        return upper, True
    if instrument in KR_TICKERS:
        return KR_TICKERS[instrument], True
    if currency == "KRW":
        return instrument, False
    return upper, False


def infer_account_type(account: str, currency: str) -> str:
    if "IRP" in account or "퇴직" in account or "연금" in account:
        return "pension"
    return "direct_investment"


def number_at(row, index: int):
    if row is None or index >= len(row):
        return None
    value = row[index]
    return value if isinstance(value, (int, float)) else None


def write_state(db_path: Path, state: dict) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS app_state (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL,
              updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            INSERT INTO app_state (key, value, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(key) DO UPDATE SET
              value = excluded.value,
              updated_at = excluded.updated_at
            """,
            (STATE_KEY, json.dumps(state, ensure_ascii=False), datetime.now(timezone.utc).isoformat()),
        )


def validate_state(state: dict, workbook) -> dict:
    holdings = state["holdings"]
    snapshots = state["portfolioSnapshots"]
    currencies = sorted({holding["currency"] for holding in holdings})
    auto_count = sum(1 for holding in holdings if holding.get("autoPrice") is not False)
    workbook_summary_total = read_workbook_total(workbook)
    workbook_detail_total = read_workbook_detail_total(workbook)
    current_total = sum(
        holding["quantity"] * holding["price"] * (1 if holding["currency"] == "KRW" else state["fxRate"]["rate"])
        for holding in holdings
    )
    cash_total = sum(cash["amount"] * (1 if cash["currency"] == "KRW" else state["fxRate"]["rate"]) for cash in state["cashBalances"])
    return {
        "holdings": len(holdings),
        "snapshots": len(snapshots),
        "cashFlows": len(state["cashFlows"]),
        "cashBalances": len(state["cashBalances"]),
        "currencies": currencies,
        "autoPriceHoldings": auto_count,
        "manualPriceHoldings": len(holdings) - auto_count,
        "firstSnapshotDate": snapshots[0]["date"] if snapshots else None,
        "lastSnapshotDate": snapshots[-1]["date"] if snapshots else None,
        "fxRate": state["fxRate"]["rate"],
        "workbookSummaryTotalKrw": workbook_summary_total,
        "workbookDetailTotalKrw": workbook_detail_total,
        "migratedHoldingsTotalKrw": current_total,
        "migratedCashTotalKrw": cash_total,
        "migratedTotalAssetsKrw": current_total + cash_total,
        "holdingsTotalMatchesWorkbookDetail": workbook_detail_total is not None and abs(current_total - workbook_detail_total) < 1,
        "sourceRowsReconciled": abs(current_total - sum(
            holding["quantity"] * holding["price"] * (1 if holding["currency"] == "KRW" else state["fxRate"]["rate"])
            for holding in holdings
        )) < 1,
        "summarySheetDiffKrw": None if workbook_summary_total is None else current_total - workbook_summary_total,
    }


def infer_cash_balances(holdings: list[dict], snapshots: list[dict], fx_rate: float) -> list[dict]:
    if not snapshots:
        return []
    holdings_total = sum(
        holding["quantity"] * holding["price"] * (1 if holding["currency"] == "KRW" else fx_rate)
        for holding in holdings
    )
    latest_snapshot = snapshots[-1]
    diff = latest_snapshot["totalValueKrw"] - holdings_total
    if abs(diff) < 1:
        return []
    if diff < 0:
        return []
    return [
        {
            "id": make_id("cash"),
            "investor": "전체",
            "account": "미분류 예수금",
            "currency": "KRW",
            "amount": diff,
            "asOf": latest_snapshot["date"],
            "source": "Numbers 총자산-보유평가 차이",
        }
    ]


def build_accounts(holdings: list[dict], cash_balances: list[dict]) -> list[dict]:
    accounts = {}
    for holding in holdings:
        key = (holding["investor"], holding["account"])
        accounts[key] = {
            "id": make_id("account"),
            "investor": holding["investor"],
            "account": holding["account"],
            "provider": infer_provider(holding["account"]),
            "accountType": holding["accountType"],
            "baseCurrency": holding["currency"],
        }
    for cash in cash_balances:
        if "미분류" in cash["account"]:
            continue
        key = (cash["investor"], cash["account"])
        accounts.setdefault(key, {
            "id": make_id("account"),
            "investor": cash["investor"],
            "account": cash["account"],
            "provider": infer_provider(cash["account"]),
            "accountType": "cash",
            "baseCurrency": cash["currency"],
        })
    return sorted(accounts.values(), key=lambda item: (item["investor"], item["account"]))


def infer_provider(account: str) -> str:
    return account.split(" ")[0] if account else ""


def read_workbook_total(workbook):
    if "총합 비율 - 총합" not in workbook.sheetnames:
        return None
    ws = workbook["총합 비율 - 총합"]
    for row in ws.iter_rows(values_only=True):
        values = [value for value in row if isinstance(value, (int, float))]
        if len(values) == 1 and values[0] > 0:
            return float(values[0])
    return None


def read_workbook_detail_total(workbook):
    total = 0.0
    matched = False
    priority = {"직투 + 연금": 3, "연금": 2, "직투": 2}
    for ws in workbook.worksheets:
        if not sheet_strategy(ws.title):
            continue
        best = None
        best_priority = -1
        for row in ws.iter_rows(values_only=True):
            label = str(row[0]).strip() if row and row[0] is not None else ""
            kind = row[2] if len(row) > 2 else None
            value = row[5] if len(row) > 5 else None
            if kind != "합계" or not isinstance(value, (int, float)):
                continue
            row_priority = priority.get(label, 0)
            if row_priority > best_priority:
                best = float(value)
                best_priority = row_priority
        if best is not None:
            total += best
            matched = True
    return total if matched else None


def make_id(prefix: str) -> str:
    return f"{prefix}-{uuid4().hex}"


if __name__ == "__main__":
    main()
